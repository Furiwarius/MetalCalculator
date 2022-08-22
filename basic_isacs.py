from openpyxl import Workbook
from openpyxl import load_workbook

lib = load_workbook(filename='lib.xlsx')
l1 = (lib['l1']) #переменная с листом с именем 'l1'


class metal (object):
    counter = 2 #это счетчик строк
                #начинается с 2, потому что 1 строка будет занята заголовками
    position_number = 1 #номер позиции у детали для ее вызова
    metal_list = [] #массив для создания имен профилей
    
    @staticmethod
    def number_check(string_1): #сравнение введенных данных с библиотекой сечений
        sent_string = string_1
        send = []
        while True:
            for n, i in enumerate(l1['A']):
                if sent_string == i.value:
                    send = [sent_string, l1['B{}'.format(n+1)].value]
                    continue
            if len(send) == 0:
                print('Введенный номер профиля отсутствует в библиотеке - {}'.format(sent_string)) 
                sent_string = input('Введите исправлненный номер профиля - ').upper() 
                if sent_string == 'ОТМЕНА':
                    return False                
            else:
                return send
    @staticmethod
    def data_change (number):
        for n, i in enumerate(metal.metal_list):
            if int(number) == n+1:
                input_1 = input("Введите нужное количество метров - ")
                i.length = metal.length_check(input_1)
                print("Выполнено")

    @staticmethod
    def length_check(string_1): #проверка введенного количества на наличие лишних символов
        sent_string = string_1
        if sent_string == 'отмена':
            return False
        else:
            while True:
                try:
                    sent_string = float(sent_string)
                    return sent_string
                except ValueError:
                    print('Введено число в сочетании с другими символами - {}'.format(sent_string))
                    sent_string = input('Введите исправленное число: ')
                    
    @staticmethod
    def delited (number):
        for n, i in enumerate(metal.metal_list):
            if int(number) == n+1:
                while True:
                    input_1 = input("Вы действительно хотите удалить этот элемент? - №{} {} длиной {} м.п. - ".format(i.position_number ,i.profile, i.length)).lower()
                    if input_1 == "да":
                        metal.metal_list.remove(i)
                        print("Выполнено")
                        return
                    elif input_1 == "нет":
                        print("Отмена")
                        return
                    else:
                        print("Введен некорректный ответ - {}".format(input_1))
        print("Такого номера позиции нет - {}".format(number))
        
        

    def __init__(self, profile):
        verification = metal.number_check(profile)
        if verification == False:
            self.__del__()
        else:
            self.profile = verification[0]                    #обязательным аргументом для создания класса является номер профиля
            print('Добавлен профиль - {}'.format(verification[0])) 
            self.unit_weight = verification[1]                      #задается после создания экземпляра
            length = metal.length_check(input('Введите длину детали: '))                           #задается после создания экземпляра
            if length == False: 
                self.__del__()
            else:
                self.length = length
                self.counter_number = metal.counter       #каждому экземпляру задается номер строки
                self.position_number = metal.position_number    #каждому экземпляру задается свой номер
                metal.position_number+=1
                metal.counter+=1
    
    def __del__(self):
        pass



def file_creation(): #функция для создания файла
    exel_file = Workbook()
    file_name = input('введите имя файла: ')
    if file_name == '': exel_file_name = 'new_exel_file.xlsx'.format(file_name) #файл будет называться new_exel_file если ничего не введено
    else: exel_file_name = '{}.xlsx'.format(file_name) #даем название создаваемому файлу
    
    exel_file_1 = exel_file.active #создаем лист exel_file_1
    title_name = input('введите имя листа: ')
    if title_name == '': exel_file_1.title = 'basic' #в файле будет называется basic если ничего не введено
    else: exel_file_1.title = '{}'.format(title_name)

    exel_file_1['A1'] = '№ п.п.'
    exel_file_1['B1'] = 'Материал'
    exel_file_1['C1'] = 'Вес 1 м.п.'
    exel_file_1['D1'] = 'Количество м.п.'
    exel_file_1['E1'] = 'Масса'
    
    # тело для записи данных в файл
    for i in metal.metal_list:
        exel_file_1['A{}'.format(i.counter_number)] = i.position_number
        exel_file_1['B{}'.format(i.counter_number)] = i.profile
        exel_file_1['C{}'.format(i.counter_number)] = i.unit_weight
        exel_file_1['D{}'.format(i.counter_number)] = i.length
        exel_file_1['E{}'.format(i.counter_number)] = '=C{}*D{}'.format(i.counter_number, i.counter_number)

    #конец записи и сохранение файла
    print('Конец записи')
    exel_file.save(filename = exel_file_name)


while True:
    print('--------------------------------------------------')
    print('0 - Выход')
    print('1 - Узнать вес нужного количеста заданного номера')
    print('2 - Узнать длину контура профиля')
    try:
        input_a = int(input('введите номер операции - '))   
    except ValueError:
        print('--------------------------------------------------')
        print('Введен некорректный номер операции - {}'.format(input_a)) 
        continue
    if input_a == 0: 
        print('Выход')
        break
    if input_a == 1:
        while True:
            print('--------------------------------------------------')
            print("№ сечения для дальнейшей работы с ним")
            print("конец - конец работы в операции 1")
            print("показ - для просмотра добавленных номеров сечений")
            print("изменить: изменать - для изменения длины детали, удалить - для удаления элемента")
            print("запись - для записи добавленных элементов в таблицу EXEL")

            input_b = input('Введите нормер сечения или другое - ')
            if input_b.lower() == 'конец': break
            elif input_b.lower() == 'показ':
                for i in metal.metal_list:
                    print('№{} {} - {} м.п., 1 м весит - {} кг, общая масса - {}'.format(i.position_number ,i.profile, i.length, i.unit_weight, i.unit_weight*i.length))
                continue


            elif input_b.lower() == 'изменить':
                print('Введите "удалить", чтобы удалить номер профиля из списка')
                print('Введите "изменить", чтобы изменить количество материала (длину)')
                print('Введите "отмена" для отмены данного действия')
                input_c = input('Введите действие - ').lower()
                
                if input_c == 'удалить':                             
                    input_d = input('Введите номер позиции удаляемого элемента - ')
                    metal.delited(input_d)
                    continue

                if input_c == 'изменить':                                
                    input_d = input('Введите номер позиции изменяемого элемента - ')
                    metal.data_change(input_d)
                    continue

                if input_c == 'отмена':
                    print('Отмена действия')
                    continue


            elif input_b.lower() == 'запись':
                file_creation()
                break
            check = metal(input_b.upper())
            metal.metal_list.append(check)
    
    if input_a == 2: #после заполнения библиотеки
        pass

        
        
